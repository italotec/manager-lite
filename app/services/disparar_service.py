import os
import sys
import json
import csv
import asyncio
import threading
import time
import uuid
import random
import string
import concurrent.futures as _cf
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from zoneinfo import ZoneInfo
_SP = ZoneInfo("America/Sao_Paulo")

import requests
from requests.adapters import HTTPAdapter

from .. import db
from ..models import DisparoJob
from ..json_store import patch_snapshot, load_user_bms

LOCK = threading.Lock()

_WEBHOOK_PROTECTED = {"PERMANENTE", "DESABILITADA", "ANALISANDO", "RESTRITA", "RETENÇÃO"}

def _flag_erro_generic_if_needed(user_id: int, waba_id: str, state: dict, msg: str) -> None:
    if not waba_id or state.get("erro_generic_marked"):
        return
    if "#135000" not in (msg or ""):
        return
    state["erro_generic_marked"] = True
    snap = (load_user_bms(user_id).get(str(waba_id).strip()) or {}).get("snapshot") or {}
    fields = {"ever_had_erro_generic": True}
    if snap.get("status_label", "") not in _WEBHOOK_PROTECTED:
        fields["status_label"] = "ERRO GENERIC"
    patch_snapshot(user_id, waba_id, **fields)


_tls = threading.local()


def _get_session() -> requests.Session:
    """One persistent HTTP session per worker thread — avoids shared pool contention."""
    if not hasattr(_tls, "session"):
        s = requests.Session()
        s.mount("https://", HTTPAdapter(pool_connections=1, pool_maxsize=1))
        _tls.session = s
    return _tls.session

# ── in-memory job state (no DB during sending) ───────────────────────────────
# {job_id: {status, total, sent, failed, skipped, last_message, stop_requested}}
_live_jobs: dict[int, dict] = {}


def get_live_state(job_id: int) -> dict | None:
    """Return in-memory state for a running job, or None if not live."""
    return _live_jobs.get(job_id)


def request_stop(job_id: int) -> bool:
    """Set stop flag in RAM. Returns True if job was live."""
    state = _live_jobs.get(job_id)
    if state:
        state["stop_requested"] = True
        return True
    return False


# ── path helpers ──────────────────────────────────────────────────────────────

def _user_base(user_id: int) -> str:
    base = os.path.join(os.getcwd(), "instance", "users", str(user_id))
    os.makedirs(base, exist_ok=True)
    return base

def csvs_dir(user_id: int) -> str:
    path = os.path.join(_user_base(user_id), "csvs")
    os.makedirs(path, exist_ok=True)
    return path

def sent_log_path(user_id: int) -> str:
    return os.path.join(_user_base(user_id), "sent_log.txt")

def disparo_log_path(user_id: int, job_id: int) -> str:
    return os.path.join(_user_base(user_id), f"disparo_log_{job_id}.jsonl")


# ── Disparou stamping (shared by _finish and restart recovery) ────────────────

def stamp_disparo_events(user_id: int, waba_id: str, sent_count: int) -> None:
    """Append a 24h disparo event for this WABA and stamp disparou_at/ultimo_disparo
    once cumulative sends in the last 24h reach 500. Pure file I/O — safe to call
    outside any running job (e.g. from restart recovery)."""
    if not waba_id or sent_count <= 0:
        return
    key = str(waba_id).strip()
    bms = load_user_bms(user_id)
    snap = (bms.get(key) or {}).get("snapshot") or {}
    events = [e for e in (snap.get("disparo_events") or []) if isinstance(e, dict)]

    now_ts = int(time.time())
    cutoff = now_ts - 86400
    events.append({"ts": now_ts, "sent": sent_count})
    events = [e for e in events if e.get("ts", 0) >= cutoff]

    total_24h = sum(e.get("sent", 0) for e in events)
    patch = {"disparo_events": events}
    if total_24h >= 500:
        patch["disparou_at"] = now_ts
        patch["ultimo_disparo"] = datetime.now(_SP).strftime("%d/%m %H:%M")
    patch_snapshot(user_id, waba_id, **patch)


def count_sent_from_log(user_id: int, job_id: int) -> int:
    """Count messages actually sent for a job by reading its real-time log.
    Used to recover the Disparou stamp when a job's process died before _finish ran."""
    path = disparo_log_path(user_id, job_id)
    if not os.path.exists(path):
        return 0
    count = 0
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                if json.loads(line).get("status") == "sent":
                    count += 1
            except Exception:
                pass
    return count


# ── random generators ─────────────────────────────────────────────────────────

def _random_namespace() -> str:
    return str(uuid.uuid4()).replace("-", "_")

def _random_param_name(length: int = 7) -> str:
    first = random.choice(string.ascii_lowercase)
    rest = "".join(random.choices(string.ascii_lowercase + string.digits, k=length - 1))
    return first + rest


# ── CSV / XLSX helpers ────────────────────────────────────────────────────────

def iter_rows(path: str, has_header: bool = True):
    """Yield rows one at a time as dicts, WITHOUT holding the whole file in RAM.

    Supports .csv and .xlsx. Use this for counting/scanning large files — building
    a full list() of a 1M-row file costs 1-2 GB and was the source of an OOM crash.
    """
    if path.lower().endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            row_iter = ws.iter_rows(values_only=True)
            first = next(row_iter, None)
            if first is None:
                return
            if has_header:
                headers = [str(c) if c is not None else "" for c in first]
            else:
                headers = [f"Coluna {i+1}" for i in range(len(first))]
                yield dict(zip(headers, [str(v) if v is not None else "" for v in first]))
            for row in row_iter:
                yield dict(zip(headers, [str(v) if v is not None else "" for v in row]))
        finally:
            wb.close()
    else:
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            first = next(reader, None)
            if first is None:
                return
            if has_header:
                headers = first
            else:
                headers = [f"Coluna {i+1}" for i in range(len(first))]
                yield dict(zip(headers, first))
            for row in reader:
                yield dict(zip(headers, row))


def _read_rows(path: str, has_header: bool = True) -> list:
    """Read all rows as list of dicts. Supports .csv and .xlsx.

    Materializes the whole file in RAM — only use when every row is genuinely
    needed at once. For counting/scanning, prefer iter_rows() (streaming).
    """
    return list(iter_rows(path, has_header=has_header))


def get_csv_columns(csv_path: str, has_header: bool = True) -> list:
    for row in iter_rows(csv_path, has_header=has_header):
        return list(row.keys())
    return []

def get_csv_preview(csv_path: str, n: int = 3, has_header: bool = True) -> list:
    out = []
    for row in iter_rows(csv_path, has_header=has_header):
        out.append(row)
        if len(out) >= n:
            break
    return out


# ── Meta API call (runs inside worker threads) ────────────────────────────────

def _send_template(phone: str, phone_number_id: str, token: str,
                   template_name: str, template_language: str,
                   parameters: list, namespace: str) -> tuple:
    """Returns (success: bool, message: str). Pure HTTP — no DB/file access."""
    api_url = f"https://graph.facebook.com/v23.0/{phone_number_id}/messages"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}",
    }

    components = []
    if parameters:
        body_params = []
        for p in parameters:
            param = {"type": "text", "text": p["value"]}
            if p.get("name"):
                param["parameter_name"] = p["name"]
            body_params.append(param)
        components.append({"type": "body", "parameters": body_params})

    payload = {
        "messaging_product": "whatsapp",
        "type": "template",
        "to": phone,
        "template": {
            "namespace": namespace,
            "name": template_name,
            "language": {"code": template_language},
            "components": components,
        },
    }

    try:
        r = _get_session().post(api_url, headers=headers, json=payload, timeout=30)
        if r.status_code == 200:
            return True, f"OK ({r.status_code})"
        return False, f"Erro {r.status_code}: {r.text[:300]}"
    except Exception as exc:
        return False, f"Exceção: {str(exc)[:300]}"


# ── async MAX mode ────────────────────────────────────────────────────────────

async def _send_template_async(session, phone, phone_number_id, token,
                               template_name, template_language, parameters, namespace):
    url = f"https://graph.facebook.com/v23.0/{phone_number_id}/messages"
    hdrs = {"Content-Type": "application/json", "Authorization": f"Bearer {token}"}
    components = []
    if parameters:
        body_params = []
        for p in parameters:
            param = {"type": "text", "text": p["value"]}
            if p.get("name"):
                param["parameter_name"] = p["name"]
            body_params.append(param)
        components.append({"type": "body", "parameters": body_params})
    payload = {
        "messaging_product": "whatsapp",
        "type": "template",
        "to": phone,
        "template": {
            "namespace": namespace,
            "name": template_name,
            "language": {"code": template_language},
            "components": components,
        },
    }
    import aiohttp as _aiohttp
    try:
        async with session.post(url, headers=hdrs, json=payload,
                                timeout=_aiohttp.ClientTimeout(total=30)) as r:
            if r.status == 200:
                return True, f"OK ({r.status})"
            text = await r.text()
            return False, f"Erro {r.status}: {text[:300]}"
    except Exception as exc:
        return False, f"Exceção: {str(exc)[:300]}"


async def _run_async_jobs(state, pending, phone_col, phone_number_id, token,
                          template_name, template_language, param_map, namespace,
                          sent_path, log_path, skip_log, user_id=0, waba_id="",
                          async_limit=500):
    import aiohttp as _aiohttp
    sem = asyncio.Semaphore(async_limit)
    # Collect results in memory — no file I/O inside coroutines (would block event loop)
    results = []  # list of (phone, success, msg, ts)

    connector = _aiohttp.TCPConnector(limit=async_limit, limit_per_host=async_limit)
    async with _aiohttp.ClientSession(connector=connector) as session:
        async def _task(row):
            async with sem:
                if state["stop_requested"]:
                    return
                phone = str(row.get(phone_col, "")).strip()
                if not phone:
                    state["skipped"] += 1
                    return
                params = [
                    {"name": pm.get("name", ""),
                     "value": str(row.get(pm.get("column", ""), "")).strip()}
                    for pm in param_map
                ]
                success, msg = await _send_template_async(
                    session, phone, phone_number_id, token,
                    template_name, template_language, params, namespace)
                ts = datetime.now(_SP).strftime("%H:%M:%S")
                results.append((phone, success, msg, ts))
                # update in-memory counters (asyncio is single-threaded — no races)
                if success:
                    state["sent"] += 1
                else:
                    state["failed"] += 1
                    _flag_erro_generic_if_needed(user_id, waba_id, state, msg)
                state["last_message"] = f"{'✓' if success else '✗'} {phone}: {msg}"

        await asyncio.gather(*[_task(row) for row in pending])

    # Batch write all logs after all requests complete — use LOCK so concurrent
    # multi-BM child jobs don't interleave writes to sent_path.
    if results:
        if not skip_log:
            sent_phones = [p for p, ok, _, _ in results if ok]
            if sent_phones:
                with LOCK:
                    with open(sent_path, "a", encoding="utf-8") as sf:
                        sf.write("\n".join(sent_phones) + "\n")
        with open(log_path, "a", encoding="utf-8") as lf:
            for phone, success, msg, ts in results:
                lf.write(json.dumps({
                    "ts": ts, "phone": phone,
                    "status": "sent" if success else "failed",
                    "message": msg,
                }, ensure_ascii=False) + "\n")


# ── background orchestrator ───────────────────────────────────────────────────

def _run_disparo(app, job_id: int, user_id: int,
                 csv_path: str, phone_col: str,
                 phone_number_id: str, token: str,
                 template_name: str, template_language: str,
                 param_map: list,
                 max_workers: int = 1,
                 skip_log: bool = False,
                 waba_id: str = "",
                 has_header: bool = True,
                 max_leads: int = 0,
                 preloaded_rows: list | None = None,
                 async_limit: int = 500):
    """
    Runs in a single daemon thread (the 'orchestrator').
    All counters live in RAM (_live_jobs). DB is only written at start and end.
    """
    # Init in-memory state
    state = {
        "status": "running",
        "total": 0, "sent": 0, "failed": 0, "skipped": 0,
        "last_message": "", "stop_requested": False,
    }
    _live_jobs[job_id] = state

    def _finish(status: str, msg: str):
        state["status"] = status
        state["last_message"] = msg
        # Stamp the Disparou check FIRST — it's the user-facing result and must never
        # be skipped because of a transient DB lock on the job-history write below.
        if waba_id and status in ("done", "stopped") and state["sent"] > 0 and not skip_log:
            try:
                stamp_disparo_events(user_id, waba_id, state["sent"])
            except Exception:
                pass
        # Single DB write at the end (job history only). A failure here (e.g. SQLite
        # "database is locked" under heavy parallel batch load) is non-fatal.
        try:
            with app.app_context():
                job = db.session.get(DisparoJob, job_id)
                if job:
                    job.status = status
                    job.total = state["total"]
                    job.sent = state["sent"]
                    job.failed = state["failed"]
                    job.skipped = state["skipped"]
                    job.last_message = msg
                    db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
        _live_jobs.pop(job_id, None)

    namespace = _random_namespace()
    sent_path = sent_log_path(user_id)
    log_path = disparo_log_path(user_id, job_id)

    # Load already-sent set. This is the dedup guard: any phone already present in
    # sent_log.txt is NEVER re-sent (unless skip_log is explicitly on, which means
    # "send to everyone, don't dedup, don't log").
    already_sent: set = set()
    if not skip_log and os.path.exists(sent_path):
        with open(sent_path, "r", encoding="utf-8") as f:
            already_sent = {ln.strip() for ln in f if ln.strip()}

    # Build `pending` in a SINGLE streaming pass. We never materialize the full file
    # and the filtered list at the same time (the old approach held both → OOM on
    # large files). preloaded_rows (multi-BM) is already an in-memory list.
    pending: list = []
    total = 0
    col_checked = False
    try:
        source = preloaded_rows if preloaded_rows is not None else iter_rows(csv_path, has_header=has_header)
        for row in source:
            total += 1
            if not col_checked:
                if phone_col not in row:
                    _finish("error", f"Coluna '{phone_col}' não encontrada no CSV.")
                    return
                col_checked = True
            # max_leads caps how many we SEND; keep counting total for the progress bar
            if max_leads > 0 and len(pending) >= max_leads:
                continue
            phone = str(row.get(phone_col, "")).strip()
            # ── dedup: skip leads already in sent_log.txt ──
            if not skip_log and phone in already_sent:
                continue
            pending.append(row)
    except Exception as exc:
        _finish("error", f"Erro ao ler arquivo: {exc}")
        return

    state["total"] = total
    state["skipped"] = total - len(pending)

    if not pending:
        _finish("done", f"Concluído — Enviados: 0  |  Falhas: 0  |  Pulados: {state['skipped']}")
        return

    def _append_log(entry: dict):
        with LOCK:
            with open(log_path, "a", encoding="utf-8") as lf:
                lf.write(json.dumps(entry, ensure_ascii=False) + "\n")

    def _worker(row: dict):
        """Pure HTTP worker — returns (phone, success, msg)."""
        phone = str(row.get(phone_col, "")).strip()
        if not phone:
            return phone, None, "telefone vazio"
        params = [
            {"name": pm.get("name", ""),
             "value": str(row.get(pm.get("column", ""), "")).strip()}
            for pm in param_map
        ]
        success, msg = _send_template(
            phone, phone_number_id, token, template_name, template_language, params, namespace
        )
        return phone, success, msg

    # ── main pool loop ─────────────────────────────────────────────────
    try:
        if max_workers == 0:
            # MAX mode: async I/O via aiohttp — 500 concurrent requests, single OS thread
            # Use SelectorEventLoop on Windows (more compatible with aiohttp than ProactorEventLoop)
            if sys.platform == "win32":
                _loop = asyncio.SelectorEventLoop()
                asyncio.set_event_loop(_loop)
                try:
                    _loop.run_until_complete(_run_async_jobs(
                        state, pending, phone_col, phone_number_id, token,
                        template_name, template_language, param_map, namespace,
                        sent_path, log_path, skip_log, user_id, waba_id,
                        async_limit,
                    ))
                finally:
                    _loop.close()
                    asyncio.set_event_loop(None)
            else:
                asyncio.run(_run_async_jobs(
                    state, pending, phone_col, phone_number_id, token,
                    template_name, template_language, param_map, namespace,
                    sent_path, log_path, skip_log, user_id, waba_id,
                    async_limit,
                ))
            if state["stop_requested"]:
                _finish("stopped", "Envio interrompido pelo usuário.")
                return
        else:
            # Sliding-window submission: never hold more than WINDOW futures in RAM
            # at once. Submitting all rows up-front (old approach) created 1M Future
            # objects + 1M row dicts simultaneously → 1-2 GB for large CSV files.
            WINDOW = max(max_workers * 4, 200)

            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                row_iter = iter(pending)
                live: dict = {}

                # Seed the initial window
                for row in row_iter:
                    if len(live) >= WINDOW:
                        break
                    live[executor.submit(_worker, row)] = row

                while live:
                    if state["stop_requested"]:
                        for f in list(live):
                            f.cancel()
                        _finish("stopped", "Envio interrompido pelo usuário.")
                        return

                    done, _ = _cf.wait(live, return_when=_cf.FIRST_COMPLETED)
                    for future in done:
                        live.pop(future)

                        try:
                            phone, success, msg = future.result()
                        except Exception as exc:
                            phone, success, msg = "?", False, str(exc)

                        if success is None:           # empty phone
                            state["skipped"] += 1
                        elif success:
                            state["sent"] += 1
                            if not skip_log:
                                with LOCK:
                                    with open(sent_path, "a", encoding="utf-8") as sf:
                                        sf.write(phone + "\n")
                        else:
                            state["failed"] += 1
                            _flag_erro_generic_if_needed(user_id, waba_id, state, msg)

                        _append_log({
                            "ts":      datetime.now(_SP).strftime("%H:%M:%S"),
                            "phone":   phone,
                            "status":  "sent" if success else "failed",
                            "message": msg,
                        })
                        state["last_message"] = f"{'✓' if success else '✗'} {phone}: {msg}"

                        # Pull in the next row to refill the window
                        next_row = next(row_iter, None)
                        if next_row is not None:
                            live[executor.submit(_worker, next_row)] = next_row

    except Exception as exc:
        _finish("error", f"Erro inesperado: {str(exc)[:300]}")
        return

    _finish(
        "done",
        f"Concluído — Enviados: {state['sent']}  |  Falhas: {state['failed']}  |  Pulados: {state['skipped']}",
    )


# ── public API ────────────────────────────────────────────────────────────────

def start_disparo_job(app, user_id: int, csv_filename: str,
                      phone_col: str, phone_number_id: str, token: str,
                      template_name: str, template_language: str,
                      param_map: list,
                      max_workers: int = 1,
                      skip_log: bool = False,
                      waba_id: str = "",
                      has_header: bool = True,
                      max_leads: int = 0,
                      preloaded_rows: list | None = None,
                      async_limit: int = 500) -> int:
    csv_path = os.path.join(csvs_dir(user_id), csv_filename)

    with app.app_context():
        job = DisparoJob(
            user_id=user_id, status="queued",
            waba_id=str(waba_id or "").strip(), skip_log=bool(skip_log),
        )
        db.session.add(job)
        db.session.commit()
        job_id = job.id

    t = threading.Thread(
        target=_run_disparo,
        args=(app, job_id, user_id, csv_path, phone_col,
              phone_number_id, token, template_name, template_language,
              param_map, max_workers, skip_log, waba_id, has_header, max_leads,
              preloaded_rows, async_limit),
        daemon=True,
    )
    t.start()
    return job_id
